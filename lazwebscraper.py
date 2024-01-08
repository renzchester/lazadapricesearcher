from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from tkinter import *
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
import ttkbootstrap as tb
import time
import sys

############################################## WEBSCRAPING ##############################################

root = tb.Window(themename = 'minty')

root.title('Lazada Webscraper')
root.geometry('360x225')

# BS4 prep part

wb = Workbook()
worksheet = wb['Sheet']

#Scraping product information in first page

product_name_list = []
product_price_list = []
product_link_list = []

def find_product():
    url = f'https://www.lazada.com.ph/tag/{search_entry.get().replace(' ', '-').lower()}'

    global browser
    if browser_combobox.get() == 'Chrome':
        browser = webdriver.Chrome()
    elif browser_combobox.get() == 'Edge':
        browser = webdriver.Edge()
    elif browser_combobox.get() == 'Firefox':
        browser = webdriver.Firefox()

    browser.get(url)
    html = browser.page_source
    soup = BeautifulSoup(html, features = 'html.parser')

    products = soup.find_all('div', class_ = 'Bm3ON')

    for product in products:
        product_name = product.find('div', class_ = 'RfADt')
        product_price = product.find('span', class_ = 'ooOxS')
        product_link = product.a['href']

        if search_entry.get().replace('-', ' ').lower() in product_name.text.lower():
            product_name_list.append(product_name.text)
            product_price_list.append(product_price.text.replace('₱', ''))
            product_link_list.append(product_link)

#Scraping product information in succeeding pages

def find_product_succeeding_pages():
 
    for i in range(int(page_entry.get())):
        if i + 1 == 1:
            continue
        elif i + 1 > int(page_entry.get()):
            break
        else:
            browser.get(f'https://www.lazada.com.ph/tag/{search_entry.get()}?page={i+1}')
            html = browser.page_source
            soup = BeautifulSoup(html, features = 'html.parser')
            find_product()

def populate_sheet():
    for row, name in zip(worksheet.iter_rows(min_row = 1, max_row = len(product_name_list)), product_name_list):
        for cell in row:
            cell.value = name

    for row, price in zip(worksheet.iter_rows(min_row = 1, max_row = len(product_name_list), min_col = 2, max_col = 2), product_price_list):
        for cell in row:
            cell.value = price

    for row, link in zip(worksheet.iter_rows(min_row = 1, max_row = len(product_name_list), min_col = 3, max_col = 3), product_link_list):
        for cell in row:
            cell.value = link

def generate_sheet():
    times_run = 0
    
    try:
        if isinstance(int(page_entry.get()), int):
            pass
        if isinstance(int(interval_integer_entry.get()), int):
            pass
    except Exception as e:
        mb = Messagebox.ok('"# of pages to search" or "Generate every __" is not a numeral!', 'Error!')

    else:
        while True:
            find_product()
            find_product_succeeding_pages()
            populate_sheet()
            times_run += 1

            wb.save(f'{search_entry.get()}-prices-links-{times_run}.xlsx')
            
            product_name_list.clear()
            product_price_list.clear()
            product_link_list.clear()
            
            if interval_unit_entry.get() == 'minute/s':
                print(f'Waiting {interval_integer_entry.get()} minute/s...')
                time.sleep(int(interval_integer_entry.get()) * 60)

            elif interval_unit_entry.get() == 'hour/s':
                print(f'Waiting {interval_integer_entry.get()} hour/s...')
                time.sleep(int(interval_integer_entry.get()) * 60 * 60)

            elif interval_unit_entry.get() == 'day/s':
                print(f'Waiting {interval_integer_entry.get()} day/s...')
                time.sleep(int(interval_integer_entry.get()) * 60 * 60 * 24)

############################################## GUI ##############################################

#Create label search
search_label = tb.Label(text = 'Search:', font=('Calibri', 14), bootstyle = 'primary')
search_label.place(x = 105, y = 10)

#Create entry field for search
search_entry = tb.Entry(bootstyle = 'primary')
search_entry.place(x = 170, y = 10)

#Create label # of pages
page_label = tb.Label(text = '# of pages to search:', font=('Calibri', 14), bootstyle = 'primary')
page_label.place(x = 5, y = 50)

#Create entry field for # of pages
page_entry = tb.Entry()
page_entry.place(x = 170, y = 50)

#Create label for 'Generate every:'
interval_label = tb.Label(text = 'Generate every:', font=('Calibri', 14), bootstyle = 'primary')
interval_label.place(x = 40, y = 90)

#Create entry field for integers for generate every
interval_integer_entry = tb.Entry()
interval_integer_entry.place(x = 170, y = 90)

#Create combobox for unit (minutes,hours,days) in generate every
interval_unit_entry_list = ['minute/s', 'hour/s', 'day/s']

interval_unit_entry = tb.Combobox(bootstyle = 'success', values = interval_unit_entry_list)
interval_unit_entry.place(x = 200, y = 90)

#Create browser combobox label
browser_label = tb.Label(text = 'Choose a browser:', font = ('Calibri', 14), bootstyle = 'primary')
browser_label.place(x = 20, y = 130)

#Create combobox for browser to be used
browser_combobox_list = ['Chrome', 'Edge', 'Firefox']

browser_combobox = tb.Combobox(bootstyle = 'success', values = browser_combobox_list)
browser_combobox.place(x = 170, y = 130)

#Create button for generating sheet
generate_sheet_button = tb.Button(text = 'Generate sheet', bootstyle = 'success', command = generate_sheet)
generate_sheet_button.place(x = 125, y = 180)

root.mainloop()