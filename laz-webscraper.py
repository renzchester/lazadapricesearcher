import requests
import json
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
from selenium import webdriver

'''
Core function
1. User can input search terms and script will automatically look for it in Lazada ===== COMPLETED
2. Results will be saved in an .xlsx file ===== COMPLETED
3. File will be saved locally, but have functionality to be saved in email
4. User will specify how many pages too look for ===== COMPLETED
5. Script will have usable UI that has all of the functionalities above
6. Script will filter out any products that do not have the search keyword entered by user ===== COMPLETED

Features that must be added
- Make an option for user which browser to use; if possible, automate based on which browser is set to default in OS
- Allow user to set number of pages to be searched ===== COMPLETED
- Allow user to search for multiple keywords
'''

search = input(f'Enter item name: ') #User inputs search keywords; must add option for inputting multiple keywords

page_number = input(f'How many pages do you want to search? ') #User inputs how many pages to scrape

# BS4 prep part

url = f'https://www.lazada.com.ph/tag/{search.replace(' ', '-').lower()}'

browser = webdriver.Chrome() #Add option for user to select which browser to use

root = 'https://lazada.com.ph'

browser.get(url)
html = browser.page_source
soup = BeautifulSoup(html, features = 'html.parser')

wb = Workbook()
worksheet = wb['Sheet']

#Scraping product information in first page

product_name_list = []
product_price_list = []
product_link_list = []

def find_product():
    products = soup.find_all('div', class_ = 'Bm3ON')

    for product in products:
        product_name = product.find('div', class_ = 'RfADt')
        product_price = product.find('span', class_ = 'ooOxS')
        product_link = product.a['href']

        if search.replace('-', ' ').lower() in product_name.text.lower():
            product_name_list.append(product_name.text)
            product_price_list.append(product_price.text.replace('₱', ''))
            product_link_list.append(product_link)

        # print(f'Product name: {product_name.text}')
        # print(f'Price: {product_price.text}')
        # print(f'Product link: {product_link}\n')

#Scraping product information in succeeding pages

# succeeding_pages = soup.find_all('li', title = True)

def find_product_succeeding_pages():

    for i in range(int(page_number)):
        if i + 1 == 1:
            continue
        elif i + 1 > int(page_number):
            break
        else:
            browser.get(f'https://www.lazada.com.ph/tag/{search}?page={i+1}')
            html = browser.page_source
            soup = BeautifulSoup(html, features = 'html.parser')
            find_product()

    # for page in succeeding_pages:
    #     page_link = page.find('a', href = True)
    #     if page_link is not None:
    #         browser.get(f'{root}/{page_link['href']}')
    #         html = browser.page_source
    #         soup = BeautifulSoup(html, features = 'html.parser')

    #         find_product()

def populate_sheet():
    for row, name in zip(worksheet.iter_rows(min_row = 1, max_row = len(product_name_list)), product_name_list):
        for cell in row:
            cell.value = name

    for row, price in zip(worksheet.iter_rows(min_row = 1, max_row = len(product_name_list), min_col = 2, max_col = 2), product_price_list):
        for cell in row:
            cell.value = price

    for row, link in zip(worksheet.iter_rows(min_row = 1, max_row = len(product_name_list), min_col = 3, max_col = 3), product_link_list):
        for cell in row:
            cell.value = link

if __name__ == '__main__':
    find_product()
    find_product_succeeding_pages()
    populate_sheet()

wb.save(f'{search}-prices-links.xlsx')